# -*- coding: utf-8 -*-
"""
Created on Thu Sep 20 10:33:24 2018
This script reads the configuration excel file

@author: TMUY
"""

# import classes
import openpyxl 


## Main
def getconfigurationfile(directory, configurationfilename,
                         sheetnames, sheetorder):
    ## open excel file in directory
    configurationfile = openpyxl.load_workbook(directory + '\\' 
                                               + configurationfilename)
    ## gets each sheet
    columnordersheet = configurationfile[sheetorder]
    columnnamessheet = configurationfile[sheetnames]
    ## get config info
    ordertmp = getconfigurationdefinition(columnordersheet)
    nametmp = getconfigurationdefinition(columnnamessheet)
    return([ordertmp,nametmp])


## load the configuration options (column names and column order)
def getconfigurationdefinition(excellsheet):
    rownumber = excellsheet.max_row
    ## temp variable where info shall be stored before returned (without 
    ## the first row, which contains the column id name)
    atmp = [0] * (rownumber - 1)
    btmp = [0] * (rownumber - 1)
    ## for: skips the first row (starts in row=2) a
    for row in range(2, rownumber):
        atmp[row-2] = excellsheet['A'+str(row)].value
        btmp[row-2] = excellsheet['B'+str(row)].value
        ## for log purpouses
        print('row:' + str(row) + str(atmp[row-2]) + '|' + str(btmp[row-2]))    
    return([atmp,btmp])


## defines the index value where 'value' is find in the input element "index"
## returns -1 is value was found
def getinfoindex(configurationdefinitionobj, value, column):
    index = -1
    try:
        index = configurationdefinitionobj[column].index(value)
    except:
        print('Log: value ' + value + ' was not found and deleted')    
    return(index)





